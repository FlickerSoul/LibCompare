#
#
#   Created By @Flicker_Soul
#   Date: Jan. 26 2020
#
#   Data Structure For LibCompareApp
#
import re
from unicodedata import normalize
from functools import partial
import openpyxl
from openpyxl import styles
from main import LOGGER, show_critical_dialog, show_warning_dialog
from collections import namedtuple


NORMALIZE_PATTERN = partial(normalize, 'NFC')
NAME_PATTERN = re.compile('[^0-9a-z\u00C0-\u00D6\u00D8-\u00F6\u00F8-\u00FF\u0100-\u017F ]+')
SPACE_PATTERN = re.compile(' +')


class CellSet:
    EQUAL = 2
    AMBIGUOUS = 1
    NOT_EQUAL = 0
    FINAL_RESULT = namedtuple('final', 'cell code')

    @staticmethod
    def process_data(cell_content: str) -> str:
        cell_content = cell_content.strip().lower()             # lower case is required
        cell_content = NORMALIZE_PATTERN(cell_content)          # normalize the string
        cell_content = cell_content.replace('&', ' and ')       # change & to and
        cell_content = NAME_PATTERN.sub(' ', cell_content)      # get rid of weird symbols
        cell_content = SPACE_PATTERN.sub(' ', cell_content)     # get rid of multiple spaces
        return cell_content

    @staticmethod
    def process_id(ids: str) -> str:
        if not ids:
            return ''

        ids = str(ids).strip()
        ids = ids.replace('-', '')
        return ids

    def __init__(self, row_num, cell_content, ids=''):
        self.row_num = row_num
        self.cell_content = str(cell_content)
        self.id = CellSet.process_id(ids)
        self.processed_data = CellSet.process_data(cell_content)
        self.possible_set = None                                  # possible values, cell:result_code
        self.final = None
        self.hash_id = hash(self.cell_content + self.processed_data + self.id)

    def init_as_comparing_cell(self):
        self.possible_set = {}
        self.final = CellSet.FINAL_RESULT(ColSet.EMPTY_CELL, CellSet.NOT_EQUAL)         # Cell:result_code
        return self

    def is_eq(self, other):
        assert isinstance(other, CellSet)

        if self.id != '' and other.id != '':
            return CellSet.EQUAL if self.id == other.id else CellSet.NOT_EQUAL

        if self.cell_content == other.cell_content or self.processed_data == other.processed_data:
            return CellSet.EQUAL
        if self.processed_data.startswith('the'):
            if self.processed_data.replace('the', '', 1).strip() == other.processed_data:
                return CellSet.EQUAL
        if other.processed_data.startswith('the'):
            if other.processed_data.replace('the', '', 1).strip() == other.processed_data:
                return CellSet.EQUAL
        if self.processed_data in other.processed_data or other.processed_data in self.processed_data or \
             self.cell_content in other.cell_content or other.cell_content in self.cell_content:
            return CellSet.AMBIGUOUS
        else:
            return CellSet.NOT_EQUAL

    def compare(self, other):
        assert isinstance(other, CellSet)

        result = self.is_eq(other)
        if not result:
            return

        self.possible_set[other] = result

    def finalize(self):
        for key, value in self.possible_set.items():
            if value == CellSet.EQUAL:
                self.final = CellSet.FINAL_RESULT(key, value)
                return

            if len(key.cell_content) > len(self.final.cell.cell_content):
                self.final = CellSet.FINAL_RESULT(key, value)               # the longer the data more detailed?

    def is_not_acceptable(self):
        return not self.final.code

    def __eq__(self, other):
        return self.hash_id == other.hash_id

    def __hash__(self):
        return self.hash_id

    def __repr__(self):
        return "content: " + self.cell_content + "\n" + \
               "processed content: " + self.processed_data + '\n' + \
               'id: ' + self.id + '\n' + \
               'hash: ' + str(self.hash_id) + '\n' + \
               'possible set: ' + (str(self.possible_set) if self.possible_set else 'null') + '\n' + \
               'final result: ' + (str(self.final) if self.final else 'null') + '\n'


class ColSet:
    A_ORD = ord('a')
    COL_OFFSET = 2
    PUBLISHER_TITLE_COL = 1
    PUBLISHER_ID_COL = 2
    LIB_TITLE_COL = 3
    LIB_ID_COL = 4
    EMPTY_CELL = CellSet(-1, 'Duplicated Empty', '')

    def __init__(self, work_sheet, title_col_num, id_col_num, row_end):
        self.sheet = work_sheet
        self.title_col_num = title_col_num
        self.id_col_num = id_col_num
        self.row_end = row_end
        self.cell_arrangement = None
        self.index_bucket: list = []
        self.duplicate_set = None
        self.duplicate_list = None

    def init_as_compared_col(self):
        self.duplicate_list = []
        return self

    @staticmethod
    def get_bucket_place(string):
        where_index = ord(string) - ColSet.A_ORD
        if where_index > 25:
            where_index = 26
        return where_index

    def get_comparing_cell(self):
        for index in range(ColSet.COL_OFFSET, self.row_end + 1):
            cell = CellSet(index,
                           self.sheet.cell(row=index, column=self.title_col_num).value,
                           self.sheet.cell(row=index, column=self.id_col_num).value
                           ).init_as_comparing_cell()

            self.cell_arrangement[ColSet.get_bucket_place(cell.processed_data[0])].append(cell)

        return self

    def get_compared_cell(self):
        assert self.cell_arrangement and self.index_bucket
        for index in range(ColSet.COL_OFFSET, self.row_end + 1):
            cell = CellSet(index,
                           self.sheet.cell(row=index, column=self.title_col_num).value,
                           self.sheet.cell(row=index, column=self.id_col_num))

            self.cell_arrangement[ColSet.get_bucket_place(cell.processed_data[0])].append(cell)
            self.index_bucket.append(cell)
        return self
    
    def get_common_comparing_cell(self):
        for index in range(ColSet.COL_OFFSET, self.row_end + 1):
            cell = CellSet(index, 
                           self.sheet.cell(row=index, column=self.title_col_num).value,
                           self.sheet.cell(row=index, column=self.id_col_num).value,
                           ).init_as_comparing_cell()
            self.index_bucket.append(cell)
            
        return self
            
    def get_common_compared_cell(self):
        for index in range(ColSet.COL_OFFSET, self.row_end + 1):
            cell = CellSet(index, 
                           self.sheet.cell(row=index, column=self.title_col_num).value,
                           self.sheet.cell(row=index, column=self.id_col_num).value,
                           )
            self.index_bucket.append(cell)
            
        return self

    @staticmethod
    def compare_alpha_subset(left, right):
        for left_cell in left:
            for right_cell in right:
                left_cell.compare(right_cell)

    def compare_col_using_subset(self, other):
        for alpha_index in range(27):
            ColSet.compare_alpha_subset(self.cell_arrangement[alpha_index], other.cell_arrangement[alpha_index])

        for subset in self.cell_arrangement:
            for cell in subset:
                cell.finalize()

    def compare_col_using_index_bucket(self, other):
        if self.index_bucket and other.index_bucket:
            for element_l in self.index_bucket:
                for element_r in other.index_bucket:
                    element_l.compare(element_r)

            for element in self.index_bucket:
                element.finalize()

    def record_matched(self, ele):
        if self.index_bucket:
            self.duplicate_list.append(ele)
            LOGGER.debug('REMOVE ELEMENT: ' + ele.processed_data)

    def take_out_matched(self):
        for ele in self.duplicate_set:
            self.index_bucket.remove(ele)

    def remove_duplicate_and_notify_duplicated(self):
        self.duplicate_set = set(self.duplicate_list)
        diff_n = len(self.duplicate_list) - len(self.duplicate_set)
        if diff_n:
            for ele in self.duplicate_set:
                self.duplicate_list.remove(ele)

        return diff_n

    def duplicate_test(self, sheet_name, col):
        s = set(self.index_bucket)
        li = self.index_bucket.copy()
        diff = len(self.index_bucket) - len(s)
        LOGGER.info(str(self.title_col_num) + ' in "' + self.sheet.title + '": duplicate: ' + str(diff) + ' titles')
        if diff:
            LOGGER.warning(str(self.title_col_num) + ' in "' + self.sheet.title + '" has duplicated titles!!')
            show_warning_dialog(
                'The Column " ' + str(self.title_col_num) + ' " In The Sheet "'
                + self.sheet.title + '" has duplicated items!',
                'Check The Duplicated Titles In Remaining Titles Sheet!'
            )

            # TODO occupying a lot of resources
            index = ColSet.COL_OFFSET
            for ele in s:
                li.remove(ele)
            for ele in li:
                sheet_name.cell(row=index, column=col, value=ele.cell_content)
                index += 1

    def __iter__(self):
        if self.index_bucket:
            for ele in self.index_bucket:
                yield ele
        else:
            for li in self.cell_arrangement:
                for ele in li:
                    yield ele

    def __repr__(self):
        return "Col From Sheet: " + self.sheet.title + \
               "\ncell: \n" + \
               '\n'.join((str(sub) for sub in (self.index_bucket if self.index_bucket else self.cell_arrangement))) \
               + "."


class WorkBookWrapper:
    PUB_TITLE_COL = 'A'
    PUB_ID_COL = 'B'
    LIB_TITLE_COL = 'C'
    LIB_ID_COL = 'D'
    DEFAULT_DATA_SHEET_NAME = 'Data'
    DEFAULT_CORRECT_COLOR = 'green'
    DEFAULT_AMBIGUOUS_COLOR = 'yellow'
    DEFAULT_WRONG_COLOR = 'red'
    a_ORD = ord('a')
    z_ORD = ord('z')
    RESULT_SHEET_NAME = 'Result'
    REMAINING_SHEET_NAME = 'Remaining Titles'

    RESULT_PUB_TITLE_COL = 1
    RESULT_PUB_ID_COL = 2
    RESULT_STATUS_COl = 3
    RESULT_MATCHED_TITLE_COL = 4
    RESULT_MATCHED_ID_COL = 5
    RESULT_PUB_PRSD_TITLE_COL = 6
    RESULT_LIB_PRSD_TITLE_COL = 7

    REMAINING_LIB_TITLE_COL = 1
    REMAINING_LIB_ID_COL = 2
    DUPLICATED_TITLES_PUB_COL = 3
    DUPLICATED_TITLES_LIB_COL = 4
    MULTIPLE_REMOVE_COL = 5

    def __init__(self, work_book_file_path):
        self.work_book = openpyxl.load_workbook(work_book_file_path)
        self.data_sheet = None
        self.data_sheet_name = None
        self.pub_title_col = None
        self.pub_id_col = None
        self.lib_title_col = None
        self.lib_id_col = None
        self.is_using_no_weird_letters = False
        self.is_using_no_space = False
        self.is_using_strict_mode = False
        self.pub_end_row = None
        self.lib_end_row = None
        self.correct_color = None
        self.ambiguous_color = None
        self.wrong_color = None

    @staticmethod
    def process_col_name(name: str) -> int:
        name = name.lower()[-1:]
        if not WorkBookWrapper.all_low_letters(name):
            raise AssertionError('Not Valid Letters')
        result = 0
        for index in range(len(name)):
            result += (ord(name[index]) - ord('a') + 1) * pow(10, index)  # the col num starts from 1

        return result

    @staticmethod
    def all_low_letters(string: str) -> bool:
        for letter in string:
            index = ord(letter)
            if index < WorkBookWrapper.a_ORD or index > WorkBookWrapper.z_ORD:
                return False
        return True

    @staticmethod
    def get_color_hex(color_string: str) -> str:
        if color_string in COLOR_TABLE:
            return COLOR_TABLE[color_string]
        else:
            raise LookupError('No Such Color')

    def get_matched_color(self, index):
        if index == CellSet.EQUAL:
            return self.correct_color
        elif index == CellSet.AMBIGUOUS:
            return self.ambiguous_color
        elif index == CellSet.NOT_EQUAL:
            return self.wrong_color
        else:
            return 'FFFFFF'

    def set_pub_id_col(self, id_col):
        self.pub_id_col = WorkBookWrapper.process_col_name(id_col)
        return self

    def set_lib_id_col(self, id_col):
        self.lib_id_col = WorkBookWrapper.process_col_name(id_col)
        return self

    def set_data_sheet(self, name):
        self.data_sheet_name = name
        self.data_sheet = self.work_book[self.data_sheet_name]

    def set_pub_title_col(self, title_col):
        self.pub_title_col = WorkBookWrapper.process_col_name(title_col)

    def set_lib_title_col(self, title_col):
        self.lib_title_col = WorkBookWrapper.process_col_name(title_col)

    def set_no_weird_letter_flag(self, flag):
        self.is_using_no_weird_letters = flag

    def set_no_space_flag(self, flag):
        self.is_using_no_space = flag

    def set_strict_mode_flag(self, flag):
        self.is_using_strict_mode = flag

    def set_pub_end_row(self, row):
        self.pub_end_row = int(row)

    def set_lib_end_row(self, row):
        self.lib_end_row = int(row)

    def set_correct_color(self, color):
        self.correct_color = WorkBookWrapper.get_color_hex(color)

    def set_ambiguous_color(self, color):
        self.ambiguous_color = WorkBookWrapper.get_color_hex(color)

    def set_wrong_color(self, color):
        self.wrong_color = WorkBookWrapper.get_color_hex(color)

    def work(self) -> bool:
        try:
            LOGGER.info('Load Publisher\'s Column')
            pub_col = ColSet(self.data_sheet, self.pub_title_col, self.pub_id_col, self.pub_end_row)\
                           .get_common_comparing_cell()
            LOGGER.debug('\n' + str(pub_col))

            LOGGER.info('Load Lib\'s Column')
            lib_col = ColSet(self.data_sheet, self.lib_title_col, self.lib_id_col, self.lib_end_row)\
                           .init_as_compared_col().get_common_compared_cell()
            LOGGER.debug('\n' + str(lib_col))
        except AttributeError:
            LOGGER.debug('Empty Row Is Detected; Please Specify A Correct End Row')
            show_critical_dialog(
                'Empty Row Is Detected',
                'Please Specify A Correct End Row'
            )
            return False

        LOGGER.info('Compare Two Columns')
        # pub_col.compare_col_using_subset(lib_col)
        pub_col.compare_col_using_index_bucket(lib_col)
        LOGGER.debug('\n' + str(pub_col))

        LOGGER.info('Dump Results')
        result_sheet = self.work_book[WorkBookWrapper.RESULT_SHEET_NAME]\
                            if WorkBookWrapper.RESULT_SHEET_NAME in self.work_book.sheetnames \
                            else self.work_book.create_sheet(WorkBookWrapper.RESULT_SHEET_NAME)
        remaining_sheet = self.work_book[WorkBookWrapper.REMAINING_SHEET_NAME]\
                               if WorkBookWrapper.REMAINING_SHEET_NAME in self.work_book.sheetnames\
                               else self.work_book.create_sheet(WorkBookWrapper.REMAINING_SHEET_NAME)

        LOGGER.info('Duplicated Title Detect')

        pub_col.duplicate_test(remaining_sheet, WorkBookWrapper.DUPLICATED_TITLES_PUB_COL)
        lib_col.duplicate_test(remaining_sheet, WorkBookWrapper.DUPLICATED_TITLES_LIB_COL)

        result_counter = ColSet.COL_OFFSET
        remaining_counter = ColSet.COL_OFFSET

        for cell in pub_col:
            # publisher's title
            result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_PUB_TITLE_COL,
                                   value=cell.cell_content)
            # publisher's title's id
            result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_PUB_ID_COL,
                                   value=cell.id)
            # match status
            result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_STATUS_COl).fill = \
                                   openpyxl.styles.PatternFill(fgColor=self.get_matched_color(cell.final.code),
                                                               fill_type='solid')
            # publisher's title processed
            result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_PUB_PRSD_TITLE_COL,
                                   value=cell.processed_data)
            if cell.final.code:
                # get target title from our lib
                target = cell.final.cell
                lib_col.record_matched(target)
                # matched title
                result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_MATCHED_TITLE_COL,
                                       value=target.cell_content)
                # matched id
                result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_MATCHED_ID_COL,
                                       value=target.id)
                # lib's title processed
                result_sheet.cell(row=result_counter, column=WorkBookWrapper.RESULT_LIB_PRSD_TITLE_COL,
                                       value=target.processed_data)
            result_counter += 1

        LOGGER.debug('\n' + str(lib_col))

        duplicated_num = lib_col.remove_duplicate_and_notify_duplicated()
        if duplicated_num:
            LOGGER.warning('Multiple Remove Of The Same Item')
            show_warning_dialog(
                'Multiple Remove Of The Same Item',
                'Some Titles Are Matched Multiple Times, Check The Remaining Sheet For More Info'
            )

        for index in range(ColSet.COL_OFFSET, ColSet.COL_OFFSET + len(lib_col.duplicate_list)):
            remaining_sheet.cell(row=index, column=WorkBookWrapper.MULTIPLE_REMOVE_COL,
                                 value=lib_col.duplicate_list[index - ColSet.COL_OFFSET].cell_content)

        lib_col.take_out_matched()
        for cell in lib_col:
            if cell != ColSet.EMPTY_CELL:
                remaining_sheet.cell(row=remaining_counter, column=WorkBookWrapper.REMAINING_LIB_TITLE_COL,
                                          value=cell.cell_content)
                remaining_sheet.cell(row=remaining_counter, column=WorkBookWrapper.REMAINING_LIB_ID_COL,
                                          value=cell.id)
                remaining_counter += 1
        LOGGER.info('Result Dumped')

        return True


COLOR_TABLE = {
    "aliceblue": 'F0F8FF',
    "antiquewhite": 'FAEBD7',
    "aqua": '00FFFF',
    "aquamarine": '7FFFD4',
    "azure": 'F0FFFF',
    "beige": 'F5F5DC',
    "bisque": 'FFE4C4',
    "black": '000000',
    "blanchedalmond": 'FFEBCD',
    "blue": '0000FF',
    "blueviolet": '8A2BE2',
    "brown": 'A52A2A',
    "burlywood": 'DEB887',
    "cadetblue": '5F9EA0',
    "chartreuse": '7FFF00',
    "chocolate": 'D2691E',
    "coral": 'FF7F50',
    "cornflowerblue": '6495ED',
    "cornsilk": 'FFF8DC',
    "crimson": 'DC143C',
    "cyan": '00FFFF',
    "darkblue": '00008B',
    "darkcyan": '008B8B',
    "darkgoldenrod": 'B8860B',
    "darkgray": 'A9A9A9',
    "darkgrey": 'A9A9A9',
    "darkgreen": '006400',
    "darkkhaki": 'BDB76B',
    "darkmagenta": '8B008B',
    "darkolivegreen": '556B2F',
    "darkorange": 'FF8C00',
    "darkorchid": '9932CC',
    "darkred": '8B0000',
    "darksalmon": 'E9967A',
    "darkseagreen": '8FBC8F',
    "darkslateblue": '483D8B',
    "darkslategray": '2F4F4F',
    "darkslategrey": '2F4F4F',
    "darkturquoise": '00CED1',
    "darkviolet": '9400D3',
    "deeppink": 'FF1493',
    "deepskyblue": '00BFFF',
    "dimgray": '696969',
    "dimgrey": '696969',
    "dodgerblue": '1E90FF',
    "firebrick": 'B22222',
    "floralwhite": 'FFFAF0',
    "forestgreen": '228B22',
    "fuchsia": 'FF00FF',
    "gainsboro": 'DCDCDC',
    "ghostwhite": 'F8F8FF',
    "gold": 'FFD700',
    "goldenrod": 'DAA520',
    "gray": '808080',
    "grey": '808080',
    "green": '008000',
    "greenyellow": 'ADFF2F',
    "honeydew": 'F0FFF0',
    "hotpink": 'FF69B4',
    "indianred": 'CD5C5C',
    "indigo": '4B0082',
    "ivory": 'FFFFF0',
    "khaki": 'F0E68C',
    "lavender": 'E6E6FA',
    "lavenderblush": 'FFF0F5',
    "lawngreen": '7CFC00',
    "lemonchiffon": 'FFFACD',
    "lightblue": 'ADD8E6',
    "lightcoral": 'F08080',
    "lightcyan": 'E0FFFF',
    "lightgoldenrodyellow": 'FAFAD2',
    "lightgray": 'D3D3D3',
    "lightgrey": 'D3D3D3',
    "lightgreen": '90EE90',
    "lightpink": 'FFB6C1',
    "lightsalmon": 'FFA07A',
    "lightseagreen": '20B2AA',
    "lightskyblue": '87CEFA',
    "lightslategray": '778899',
    "lightslategrey": '778899',
    "lightsteelblue": 'B0C4DE',
    "lightyellow": 'FFFFE0',
    "lime": '00FF00',
    "limegreen": '32CD32',
    "linen": 'FAF0E6',
    "magenta": 'FF00FF',
    "maroon": '800000',
    "mediumaquamarine": '66CDAA',
    "mediumblue": '0000CD',
    "mediumorchid": 'BA55D3',
    "mediumpurple": '9370DB',
    "mediumseagreen": '3CB371',
    "mediumslateblue": '7B68EE',
    "mediumspringgreen": '00FA9A',
    "mediumturquoise": '48D1CC',
    "mediumvioletred": 'C71585',
    "midnightblue": '191970',
    "mintcream": 'F5FFFA',
    "mistyrose": 'FFE4E1',
    "moccasin": 'FFE4B5',
    "navajowhite": 'FFDEAD',
    "navy": '000080',
    "oldlace": 'FDF5E6',
    "olive": '808000',
    "olivedrab": '6B8E23',
    "orange": 'FFA500',
    "orangered": 'FF4500',
    "orchid": 'DA70D6',
    "palegoldenrod": 'EEE8AA',
    "palegreen": '98FB98',
    "paleturquoise": 'AFEEEE',
    "palevioletred": 'DB7093',
    "papayawhip": 'FFEFD5',
    "peachpuff": 'FFDAB9',
    "peru": 'CD853F',
    "pink": 'FFC0CB',
    "plum": 'DDA0DD',
    "powderblue": 'B0E0E6',
    "purple": '800080',
    "rebeccapurple": '663399',
    "red": 'FF0000',
    "rosybrown": 'BC8F8F',
    "royalblue": '4169E1',
    "saddlebrown": '8B4513',
    "salmon": 'FA8072',
    "sandybrown": 'F4A460',
    "seagreen": '2E8B57',
    "seashell": 'FFF5EE',
    "sienna": 'A0522D',
    "silver": 'C0C0C0',
    "skyblue": '87CEEB',
    "slateblue": '6A5ACD',
    "slategray": '708090',
    "slategrey": '708090',
    "snow": 'FFFAFA',
    "springgreen": '00FF7F',
    "steelblue": '4682B4',
    "tan": 'D2B48C',
    "teal": '008080',
    "thistle": 'D8BFD8',
    "tomato": 'FF6347',
    "turquoise": '40E0D0',
    "violet": 'EE82EE',
    "wheat": 'F5DEB3',
    "white": 'FFFFFF',
    "whitesmoke": 'F5F5F5',
    "yellow": 'FFFF00',
    "yellowgreen": '9ACD32'
}


def output_color_name():
    keys = COLOR_TABLE.keys()
    lk = list(keys)
    lk.sort()
    return lk


if __name__ == '__main__':
    END = 4
    work_book = openpyxl.load_workbook('test/test.xlsx')
    sheet = work_book['Data']
    d = ColSet(sheet, ColSet.PUBLISHER_TITLE_COL, ColSet.PUBLISHER_ID_COL, END).get_comparing_cell()
    print(d)
    c = ColSet(sheet, ColSet.LIB_TITLE_COL, ColSet.LIB_ID_COL, END).get_comparing_cell()
    print(c)
    d.compare_col_using_subset(c)
    print(d)

    print(output_color_name())
