import openpyxl
from openpyxl import styles
import re

NAME_PATTERN = re.compile('[^0-9a-z ]+')
SPACE_PATTERN = re.compile(' +')
temp_index = 0

FROM_TEMP_COL = 1
TO_TEMP_COL = 5

FROM_FINAL_COL = 1
FROM_CACHE_FINAL_COL = 5
TO_FINAL_COL = 6
TO_CACHE_FINAL_COL = 7
CORRESPING_FINAL_COL = 3
FILL_CELL = 2

REMAINING_COL = 1


def process_col_name(name: str) -> int :
    result = 0
    name = name.lower()[-1:]
    for index in range(len(name)):
        result += (ord(name[index]) - ord('a') + 1) * pow(10, index)      # the col num starts from 1
    
    return result

        
def get_work_book(path_to_work_book: str):
    wb = openpyxl.load_workbook(filename=path_to_work_book)
    return wb


def get_work_sheet(name: str, work_book):
    sheet = None
    try:
        sheet = work_book[name]
    except KeyError:
        print('No Such Sheet As ' + name)

    return sheet


def process_name(name: str) -> str:
    name = NAME_PATTERN.sub('', name.lower())
    name = SPACE_PATTERN.sub(' ', name)

    return name


def create_cache_sheet(workbook):
    global temp_index
    temp_index += 1
    return workbook.create_sheet('temp' + str(temp_index))


def create_final_sheet(workbook):
    return workbook.create_sheet('final result')


def create_remaining_sheet(workbook):
    return workbook.create_sheet('remaining titles')


def cache_names(current_sheet, cache_sheet, col, end_row, target_col):
    for index in range(1, end_row + 1):
        cache_sheet.cell(row=index, column=target_col,
                         value=process_name(str(current_sheet.cell(row=index, column=col).value)))


def validate_name(f: str, t: str):
    f = str(f)
    t = str(t)
    if f == t:
        return 2
    elif f in t or (t != '' and t in f):
        return 1
    else:
        return 0


def fill_cell(cell, color):
    cell.fill = openpyxl.styles.PatternFill(fgColor=color, fill_type='solid')


def cp_to_final_sheet(sheet, fs, col, end, fcol):
    for index in range(1, end+1):
        fs.cell(column=fcol, row=index, value=sheet.cell(column=col, row=index).value)


def cp_to_cell(from_sheet, fr, fc, to_sheet, tr, tc):
    to_sheet.cell(row=tr, column=tc, value=from_sheet.cell(row=fr, column=fc).value)


def gather_incompatable(original_sheet, cache_sheet, final_sheet, original_col, cache_col, final_col, cache_end):
    counter = 0
    for index in range(1, cache_end+1):
        if str(cache_sheet.cell(row=index, column=cache_col).value) == '':
            continue
        counter += 1
        final_sheet.cell(row=counter, column=final_col, value=original_sheet.cell(row=index, column=original_col).value)


def start(fromcol, tocol, work_book_path, data_sheet_name, from_end, to_end,
          auto_select_flag, only_letter_flag, no_space_flag, strict_compare_flag,
          correct_color_name, ambiguous_color_name, wrong_color_name):
    wb = get_work_book(work_book_path)
    work_sheet = get_work_sheet(data_sheet_name, wb)
    fromcol = process_col_name(fromcol)
    tocol = process_col_name(tocol)
    cache_sheet = create_cache_sheet(wb)

    cache_names(work_sheet, cache_sheet, fromcol, from_end, FROM_TEMP_COL)
    cache_names(work_sheet, cache_sheet, tocol, from_end, TO_TEMP_COL)

    final_sheet = create_final_sheet(wb)
    cp_to_final_sheet(work_sheet, final_sheet, fromcol, from_end, FROM_FINAL_COL)
    cp_to_final_sheet(work_sheet, final_sheet, tocol, to_end, TO_FINAL_COL)
    cp_to_final_sheet(cache_sheet, final_sheet, FROM_TEMP_COL, from_end, FROM_CACHE_FINAL_COL)
    cp_to_final_sheet(cache_sheet, final_sheet, TO_TEMP_COL, to_end, TO_CACHE_FINAL_COL)

    remaining_sheet = create_remaining_sheet(wb)

    # processing

    for index in range(1, from_end+1):
        flag = True
        for t in range(1, to_end+1):
            result = validate_name(cache_sheet.cell(row=index, column=FROM_TEMP_COL).value,
                                   cache_sheet.cell(row=t, column=TO_TEMP_COL).value)

            if result == 0:
                continue
            elif result == 1:
                fill_cell(final_sheet.cell(row=index, column=FILL_CELL), ambiguous_color_name)
                cp_to_cell(final_sheet, t, TO_FINAL_COL, final_sheet, index, CORRESPING_FINAL_COL)

                flag = False
                cache_sheet.cell(row=t, column=TO_TEMP_COL, value='')
                break
            elif result == 2:
                fill_cell(final_sheet.cell(row=index, column=FILL_CELL), correct_color_name)
                cp_to_cell(final_sheet, t, TO_FINAL_COL, final_sheet, index, CORRESPING_FINAL_COL)
                cache_sheet.cell(row=t, column=TO_TEMP_COL, value='')
                flag = False
                break
        if flag:
            fill_cell(final_sheet.cell(row=index, column=FILL_CELL), wrong_color_name)

    gather_incompatable(work_sheet, cache_sheet, remaining_sheet, tocol, TO_TEMP_COL, REMAINING_COL, from_end)
    wb.save('reuslt.xlsx')


if __name__ == "__main__":
    file_path = '/test/test.xlsx'
    sheet_name = "Sheet1"
    from_col = 'A'
    to_col = 'E'

    from_end_row = 1402
    to_end_row = 1572

    auto_select = False
    only_letter = True
    no_space = False
    strict_compare = False

    correct_color = '00B300'
    ambiguous_color = 'FFB733'
    wrong_color = 'CE0A0A'

    start(from_col, to_col, file_path, sheet_name, from_end_row, to_end_row,
          auto_select, only_letter, no_space, strict_compare,
          correct_color, ambiguous_color, wrong_color)

